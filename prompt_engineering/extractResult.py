import openpyxl

def extract_last_b_column_value(input_file, output_file):
    # Load the input workbook
    wb = openpyxl.load_workbook(input_file, data_only=True)
    
    # Create a new workbook for output
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(["Sheet Name", "Last B Column Value"])
    
    # Iterate through all sheets
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Find the last row with data in column B
        last_row = ws.max_row
        while last_row > 0 and ws[f"B{last_row}"].value is None:
            last_row -= 1
        
        # Get the value from column B of the last row
        last_value = ws[f"B{last_row}"].value if last_row > 0 else None
        
        # Append data to the output sheet
        out_ws.append([sheet, last_value])
    
    # Save the output workbook
    out_wb.save(output_file)

# Example usage
input_file = "chatbot_responses.xlsx"
output_file = "resultBook.xlsx"
extract_last_b_column_value(input_file, output_file)
